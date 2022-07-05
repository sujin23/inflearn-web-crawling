import openpyxl

# 1) 엑셀 불러오기
wb = openpyxl.load_workbook(r'C:\crawling\Chapter05\2022_T1_lol_Player.xlsx')

# 2) 엑셀 워크시트 선택하기
ws = wb['2022 T1']

# 3) 데이터 수정하기
ws['C1'] = '포지션'

ws['C2'] = '탑'
ws['C3'] = '정글'
ws['C4'] = '미드'
ws['C5'] = '바텀'
ws['C6'] = '서포터'

# 4) 엑셀 저장하기
wb.save(r'C:\crawling\Chapter05\2022_T1_lol_Player.xlsx')