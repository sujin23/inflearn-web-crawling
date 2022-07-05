import openpyxl

# 1) 엑셀 만들기
wb = openpyxl.Workbook()

# 2) 엑셀 워크시트 만들기
ws = wb.create_sheet('2022 T1')

# 3) 데이터 추가하기
ws['A1'] = '소환사명'
ws['B1'] = '성명'

ws['A2'] = 'Zeus'
ws['B2'] = '최우제'

ws['A3'] = 'Oner'
ws['B3'] = '문현준'

ws['A4'] = 'Faker'
ws['B4'] = '이상혁'

ws['A5'] = 'Gumayusi'
ws['B5'] = '이민형'

ws['A6'] = 'Keria'
ws['B6'] = '류민석'

ws['A1'] = '소환사명'
ws['B1'] = '성명'

# 4) 엑셀 저장하기
wb.save(r'C:\crawling\Chapter05\2022_T1_lol_Player.xlsx')